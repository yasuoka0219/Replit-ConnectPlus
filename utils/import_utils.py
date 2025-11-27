"""
CSV/Excel インポートユーティリティ
"""
import csv
import io
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

# 業界名のマッピング（一般的な別名に対応）
INDUSTRY_NAME_MAPPING = {
    # マーケティング・広告関連
    'マーケティング': '広告・メディア',
    'マーケティング・広告': '広告・メディア',
    '広告・PR': '広告・メディア',
    'PR・広報': '広告・メディア',
    '広告': '広告・メディア',
    'メディア': '広告・メディア',
    '出版・印刷': '広告・メディア',
    '出版': '広告・メディア',
    '印刷': '広告・メディア',
    # 小売関連
    '小売・流通': '小売・卸売',
    '小売業': '小売・卸売',
    '流通業': '小売・卸売',
    '卸売業': '小売・卸売',
    '小売': '小売・卸売',
    'アパレル': '小売・卸売',
    'ファッション': '小売・卸売',
    # 飲食関連
    '飲食': '飲食・宿泊',
    '飲食業': '飲食・宿泊',
    '食品': '飲食・宿泊',
    '宿泊': '飲食・宿泊',
    'ホテル': '飲食・宿泊',
    # 医療関連
    '医療・ヘルスケア': '医療・福祉',
    'ヘルスケア': '医療・福祉',
    '医療': '医療・福祉',
    '福祉': '医療・福祉',
    '製薬': '医療・福祉',
    '製薬業': '医療・福祉',
    # IT関連
    'IT': 'IT・ソフトウェア',
    'ソフトウェア': 'IT・ソフトウェア',
    'IT・通信': 'IT・ソフトウェア',
    '通信・IT': 'IT・ソフトウェア',
    '通信': 'IT・ソフトウェア',
    # 建設関連
    '建設': '建設・不動産',
    '建設業': '建設・不動産',
    '不動産': '建設・不動産',
    '建築': '建設・不動産',
    # 金融関連
    '金融': '金融・保険',
    '金融業': '金融・保険',
    '保険': '金融・保険',
    '銀行': '金融・保険',
    '証券': '金融・保険',
    # 物流関連
    '運輸・物流': '物流・運輸',
    '物流': '物流・運輸',
    '運輸': '物流・運輸',
    '運送': '物流・運輸',
    # 教育関連
    '教育': '人材・教育',
    '教育業': '人材・教育',
    '人材': '人材・教育',
    # 公共関連
    '公共サービス': '自治体・公共',
    '公共': '自治体・公共',
    '自治体': '自治体・公共',
    '官公庁': '自治体・公共',
    # エンタメ・スポーツ関連
    '健康・スポーツ': 'エンタメ・スポーツ',
    'スポーツ': 'エンタメ・スポーツ',
    'エンタメ': 'エンタメ・スポーツ',
    'エンターテイメント': 'エンタメ・スポーツ',
    '観光・レジャー': 'エンタメ・スポーツ',
    '観光': 'エンタメ・スポーツ',
    'レジャー': 'エンタメ・スポーツ',
    # エネルギー・環境関連
    '環境': 'エネルギー・インフラ',
    'エネルギー': 'エネルギー・インフラ',
    'インフラ': 'エネルギー・インフラ',
    # 専門サービス関連
    'コンサルティング': '専門サービス',
    'コンサル': '専門サービス',
    'サービス業': '専門サービス',
    '専門サービス': '専門サービス',
    # その他
    '農業': 'その他',
    '農林水産': 'その他',
}


def normalize_industry_name(industry_name):
    """業界名を正規化（マッピング辞書を使用）"""
    if not industry_name:
        return None
    
    industry_name = industry_name.strip()
    
    # マッピング辞書で変換
    if industry_name in INDUSTRY_NAME_MAPPING:
        return INDUSTRY_NAME_MAPPING[industry_name]
    
    # マッピングされていない場合はそのまま返す
    return industry_name


def parse_csv_file(file):
    """CSVファイルをパースしてデータ行のリストを返す"""
    file_content = file.read()
    
    # UTF-8でデコードを試行
    try:
        content_str = file_content.decode('utf-8-sig')  # BOMを除去
    except UnicodeDecodeError:
        # Shift-JISでデコードを試行
        try:
            content_str = file_content.decode('shift_jis')
        except UnicodeDecodeError:
            content_str = file_content.decode('utf-8', errors='ignore')
    
    file.stream.seek(0)
    
    reader = csv.DictReader(io.StringIO(content_str))
    rows = []
    for row in reader:
        # 空の行をスキップ
        if any(row.values()):
            rows.append(row)
    return rows


def parse_excel_file(file):
    """Excelファイルをパースしてデータ行のリストを返す"""
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    
    # ヘッダー行を取得
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value) if cell.value else '')
    
    # データ行を取得
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_dict = {}
        for idx, cell in enumerate(row):
            if idx < len(headers):
                value = cell.value
                # 日付型の場合は文字列に変換
                if isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif hasattr(value, 'date'):  # date型の場合
                    value = value.strftime('%Y-%m-%d')
                row_dict[headers[idx]] = str(value) if value is not None else ''
        
        # 空の行をスキップ
        if any(row_dict.values()):
            rows.append(row_dict)
    
    return rows


def validate_company_row(row, row_num, industry_categories=None):
    """企業データの行をバリデーション"""
    errors = []
    
    # 必須フィールド
    if not row.get('企業名') and not row.get('name'):
        errors.append(f"行{row_num}: 企業名は必須です")
    
    # 業界のバリデーション（オプション）
    industry = row.get('業界') or row.get('industry', '')
    if industry:
        # 業界名を正規化
        normalized_industry = normalize_industry_name(industry)
        
        # 正規化後の業界名でバリデーション
        if industry_categories:
            # マッピングが適用されたかチェック
            was_mapped = (industry != normalized_industry and normalized_industry in INDUSTRY_NAME_MAPPING.values())
            
            if normalized_industry in industry_categories:
                # 正規化後の業界名が有効な場合、正規化された業界名をrowに設定
                row['_normalized_industry'] = normalized_industry
                # マッピングが適用された場合は警告として記録（エラーにはしない）
                if was_mapped:
                    row['_industry_mapped'] = True
            else:
                # マッピングできない場合のみエラー
                error_msg = f"行{row_num}: 無効な業界 '{industry}' です。"
                error_msg += f" 有効な業界: {', '.join(industry_categories[:5])}..."
                errors.append(error_msg)
    
    # 温度感スコアのバリデーション（オプション）
    heat_score = row.get('温度感スコア') or row.get('heat_score', '')
    if heat_score:
        try:
            score = int(heat_score)
            if not (1 <= score <= 5):
                errors.append(f"行{row_num}: 温度感スコアは1-5の範囲で指定してください")
        except ValueError:
            errors.append(f"行{row_num}: 温度感スコアは数値で指定してください")
    
    return errors


def validate_deal_row(row, row_num, companies_dict):
    """案件データの行をバリデーション"""
    errors = []
    
    # 必須フィールド
    if not row.get('案件名') and not row.get('title'):
        errors.append(f"行{row_num}: 案件名は必須です")
    
    company_name = row.get('企業名') or row.get('company_name', '')
    if not company_name:
        errors.append(f"行{row_num}: 企業名は必須です")
    elif company_name not in companies_dict:
        errors.append(f"行{row_num}: 企業 '{company_name}' が見つかりません（先に企業を登録してください）")
    
    # ステージのバリデーション（オプション）
    stage = row.get('ステージ') or row.get('stage', '')
    if stage:
        valid_stages = ['初回接触', '提案', '見積', '交渉', '成約']
        if stage not in valid_stages:
            errors.append(f"行{row_num}: 無効なステージ '{stage}' です")
    
    # ステータスのバリデーション（オプション）
    status = row.get('ステータス') or row.get('status', '')
    if status:
        valid_statuses = ['OPEN', 'WON', 'LOST', '進行中', '受注', '失注', '成約']
        if status not in valid_statuses:
            # ステータス名を英語に変換
            status_map = {'進行中': 'OPEN', '受注': 'WON', '失注': 'LOST', '成約': 'WON'}
            if status in status_map:
                row['status'] = status_map[status]
            else:
                errors.append(f"行{row_num}: 無効なステータス '{status}' です")
    
    # 金額のバリデーション（オプション）
    amount = row.get('金額') or row.get('amount', '')
    if amount:
        try:
            float(amount)
        except ValueError:
            errors.append(f"行{row_num}: 金額は数値で指定してください")
    
    return errors

