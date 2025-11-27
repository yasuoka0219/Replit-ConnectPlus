"""
CSV/Excel エクスポート・インポートユーティリティ
"""
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def export_companies_to_csv(companies):
    """企業データをCSV形式でエクスポート"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # ヘッダー行
    headers = [
        'ID', '企業名', '業界', '所在地', '本社所在地', '従業員数', 
        'ウェブサイト', '温度感スコア', 'タグ', '最終接触日', '次回アクション予定日',
        'メモ', 'ニーズ', '現状KPI', '作成日'
    ]
    writer.writerow(headers)
    
    # データ行
    for company in companies:
        writer.writerow([
            company.id,
            company.name,
            company.industry or '',
            company.location or '',
            company.hq_location or '',
            company.employee_size or '',
            company.website or '',
            company.heat_score or '',
            company.tags or '',
            company.last_contacted_at.strftime('%Y-%m-%d %H:%M:%S') if company.last_contacted_at else '',
            company.next_action_at.strftime('%Y-%m-%d %H:%M:%S') if company.next_action_at else '',
            company.memo or '',
            company.needs or '',
            company.kpi_current or '',
            company.created_at.strftime('%Y-%m-%d %H:%M:%S') if company.created_at else ''
        ])
    
    return output.getvalue()


def export_companies_to_excel(companies):
    """企業データをExcel形式でエクスポート"""
    wb = Workbook()
    ws = wb.active
    ws.title = "企業一覧"
    
    # ヘッダー行
    headers = [
        'ID', '企業名', '業界', '所在地', '本社所在地', '従業員数', 
        'ウェブサイト', '温度感スコア', 'タグ', '最終接触日', '次回アクション予定日',
        'メモ', 'ニーズ', '現状KPI', '作成日'
    ]
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行
    for row_idx, company in enumerate(companies, 2):
        ws.cell(row=row_idx, column=1, value=company.id)
        ws.cell(row=row_idx, column=2, value=company.name)
        ws.cell(row=row_idx, column=3, value=company.industry or '')
        ws.cell(row=row_idx, column=4, value=company.location or '')
        ws.cell(row=row_idx, column=5, value=company.hq_location or '')
        ws.cell(row=row_idx, column=6, value=company.employee_size or '')
        ws.cell(row=row_idx, column=7, value=company.website or '')
        ws.cell(row=row_idx, column=8, value=company.heat_score or '')
        ws.cell(row=row_idx, column=9, value=company.tags or '')
        ws.cell(row=row_idx, column=10, value=company.last_contacted_at.strftime('%Y-%m-%d %H:%M:%S') if company.last_contacted_at else '')
        ws.cell(row=row_idx, column=11, value=company.next_action_at.strftime('%Y-%m-%d %H:%M:%S') if company.next_action_at else '')
        ws.cell(row=row_idx, column=12, value=company.memo or '')
        ws.cell(row=row_idx, column=13, value=company.needs or '')
        ws.cell(row=row_idx, column=14, value=company.kpi_current or '')
        ws.cell(row=row_idx, column=15, value=company.created_at.strftime('%Y-%m-%d %H:%M:%S') if company.created_at else '')
    
    # 列幅の自動調整
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # ファイルをメモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_deals_to_csv(deals):
    """案件データをCSV形式でエクスポート"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # ヘッダー行
    headers = [
        'ID', '企業名', '案件名', 'ステージ', '金額', 'ステータス', '担当者', 
        '温度感スコア', 'アポイント日', '次回アクション', '受注理由カテゴリ', 
        '受注理由詳細', '失注理由カテゴリ', '失注理由詳細', 'クローズ日', 
        '作成日', 'メモ', '議事録'
    ]
    writer.writerow(headers)
    
    # データ行
    for deal in deals:
        writer.writerow([
            deal.id,
            deal.company.name if deal.company else '',
            deal.title,
            deal.stage,
            deal.amount or 0,
            deal.status,
            deal.get_assignee_name() or deal.assignee or '',
            deal.heat_score or '',
            deal.appointment_date.strftime('%Y-%m-%d') if deal.appointment_date else '',
            deal.next_action or '',
            deal.win_reason_category or '',
            deal.win_reason_detail or '',
            deal.lost_reason_category or '',
            deal.lost_reason_detail or '',
            deal.closed_at.strftime('%Y-%m-%d %H:%M:%S') if deal.closed_at else '',
            deal.created_at.strftime('%Y-%m-%d %H:%M:%S') if deal.created_at else '',
            deal.note or '',
            deal.meeting_minutes or ''
        ])
    
    return output.getvalue()


def export_deals_to_excel(deals):
    """案件データをExcel形式でエクスポート"""
    wb = Workbook()
    ws = wb.active
    ws.title = "案件一覧"
    
    # ヘッダー行
    headers = [
        'ID', '企業名', '案件名', 'ステージ', '金額', 'ステータス', '担当者', 
        '温度感スコア', 'アポイント日', '次回アクション', '受注理由カテゴリ', 
        '受注理由詳細', '失注理由カテゴリ', '失注理由詳細', 'クローズ日', 
        '作成日', 'メモ', '議事録'
    ]
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行
    for row_idx, deal in enumerate(deals, 2):
        ws.cell(row=row_idx, column=1, value=deal.id)
        ws.cell(row=row_idx, column=2, value=deal.company.name if deal.company else '')
        ws.cell(row=row_idx, column=3, value=deal.title)
        ws.cell(row=row_idx, column=4, value=deal.stage)
        ws.cell(row=row_idx, column=5, value=deal.amount or 0)
        ws.cell(row=row_idx, column=6, value=deal.status)
        ws.cell(row=row_idx, column=7, value=deal.get_assignee_name() or deal.assignee or '')
        ws.cell(row=row_idx, column=8, value=deal.heat_score or '')
        ws.cell(row=row_idx, column=9, value=deal.appointment_date.strftime('%Y-%m-%d') if deal.appointment_date else '')
        ws.cell(row=row_idx, column=10, value=deal.next_action or '')
        ws.cell(row=row_idx, column=11, value=deal.win_reason_category or '')
        ws.cell(row=row_idx, column=12, value=deal.win_reason_detail or '')
        ws.cell(row=row_idx, column=13, value=deal.lost_reason_category or '')
        ws.cell(row=row_idx, column=14, value=deal.lost_reason_detail or '')
        ws.cell(row=row_idx, column=15, value=deal.closed_at.strftime('%Y-%m-%d %H:%M:%S') if deal.closed_at else '')
        ws.cell(row=row_idx, column=16, value=deal.created_at.strftime('%Y-%m-%d %H:%M:%S') if deal.created_at else '')
        ws.cell(row=row_idx, column=17, value=deal.note or '')
        ws.cell(row=row_idx, column=18, value=deal.meeting_minutes or '')
    
    # 列幅の自動調整
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # ファイルをメモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_activities_to_csv(activities):
    """活動履歴データをCSV形式でエクスポート"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # ヘッダー行
    headers = [
        'ID', '企業名', 'ユーザー名', '案件名', '活動タイプ', 'タイトル', 
        '内容', '実施日時', '作成日時'
    ]
    writer.writerow(headers)
    
    # データ行
    for activity in activities:
        writer.writerow([
            activity.id,
            activity.company.name if activity.company else '',
            activity.user.name if activity.user else '',
            activity.deal.title if activity.deal else '',
            activity.type,
            activity.title,
            activity.body or '',
            activity.happened_at.strftime('%Y-%m-%d %H:%M:%S') if activity.happened_at else '',
            activity.created_at.strftime('%Y-%m-%d %H:%M:%S') if activity.created_at else ''
        ])
    
    return output.getvalue()


def export_activities_to_excel(activities):
    """活動履歴データをExcel形式でエクスポート"""
    wb = Workbook()
    ws = wb.active
    ws.title = "活動履歴"
    
    # ヘッダー行
    headers = [
        'ID', '企業名', 'ユーザー名', '案件名', '活動タイプ', 'タイトル', 
        '内容', '実施日時', '作成日時'
    ]
    
    # ヘッダースタイル
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行
    for row_idx, activity in enumerate(activities, 2):
        ws.cell(row=row_idx, column=1, value=activity.id)
        ws.cell(row=row_idx, column=2, value=activity.company.name if activity.company else '')
        ws.cell(row=row_idx, column=3, value=activity.user.name if activity.user else '')
        ws.cell(row=row_idx, column=4, value=activity.deal.title if activity.deal else '')
        ws.cell(row=row_idx, column=5, value=activity.type)
        ws.cell(row=row_idx, column=6, value=activity.title)
        ws.cell(row=row_idx, column=7, value=activity.body or '')
        ws.cell(row=row_idx, column=8, value=activity.happened_at.strftime('%Y-%m-%d %H:%M:%S') if activity.happened_at else '')
        ws.cell(row=row_idx, column=9, value=activity.created_at.strftime('%Y-%m-%d %H:%M:%S') if activity.created_at else '')
    
    # 列幅の自動調整
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # ファイルをメモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


